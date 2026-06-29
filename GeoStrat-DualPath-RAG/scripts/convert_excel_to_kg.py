"""
将岩石地层Excel (521条) 转换为知识图谱三元组
数据源: E:\地质文本空间化平台\岩石地层-0312地大武汉.xlsx
输出: kb_storage/graphs/experiment_kb.graphml
"""
import json, os, sys, re
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import openpyxl
import networkx as nx
from collections import Counter

# ═══════════════════════════════════════════════════════════
# 本体定义
# ═══════════════════════════════════════════════════════════

ENTITY_TYPES = {
    "GROUP": "群", "FORMATION": "组", "MEMBER": "段",
    "LITHOLOGY": "岩石类型", "GEOLOGICAL_TIME": "地质年代",
    "LOCATION": "地理位置", "FOSSIL": "化石",
    "THICKNESS": "厚度", "CONTACT_RELATIONSHIP": "接触关系"
}

# 岩石名称正则 (常见岩石类型)
ROCK_PATTERNS = [
    r'(?:[一-鿿]{0,3}(?:岩|石|矿|土|砾|砂|泥|脉|体))',
]
# 更具体的岩石类型提取
SPECIFIC_ROCKS = [
    # 变质岩
    r'[^\s,，、；;。]{1,3}(?:片岩|麻岩|变粒岩|角闪岩|大理岩|石英岩|板岩|千枚岩|混合岩|糜棱岩|碎裂岩|矽卡岩)',
    # 沉积岩
    r'[^\s,，、；;。]{1,3}(?:砂岩|砾岩|泥岩|页岩|灰岩|白云岩|硅质岩|磷块岩|煤层|铝土矿|铁质岩|锰质岩)',
    # 火山岩/岩浆岩
    r'[^\s,，、；;。]{1,3}(?:花岗岩|闪长岩|辉长岩|橄榄岩|玄武岩|安山岩|流纹岩|凝灰岩|英安岩|粗面岩|响岩|正长岩|二长岩|斑岩|玢岩|角斑岩|细碧岩|蛇绿岩|科马提岩)',
    # 火山碎屑岩
    r'(?:火山角砾岩|集块岩|凝灰质[^\s,，、；;。]{0,3})',
    # 松散沉积物
    r'(?:[^\s,，、；;。]{0,2}(?:砾石|砂|粉砂|黏土|淤泥|泥炭))',
]

# 常见岩石名称词汇
ROCK_NAMES = {
    # 变质岩
    '片岩', '片麻岩', '变粒岩', '角闪岩', '大理岩', '石英岩', '板岩', '千枚岩',
    '混合岩', '糜棱岩', '碎裂岩', '矽卡岩', '角岩', '麻粒岩', '榴辉岩',
    # 沉积岩
    '砂岩', '砾岩', '泥岩', '页岩', '灰岩', '白云岩', '硅质岩', '磷块岩',
    '煤层', '铝土矿', '铁质岩', '锰质岩', '泥灰岩', '蒸发岩', '浊积岩',
    # 火山岩/岩浆岩
    '花岗岩', '闪长岩', '辉长岩', '橄榄岩', '玄武岩', '安山岩', '流纹岩',
    '凝灰岩', '英安岩', '粗面岩', '响岩', '正长岩', '二长岩', '斑岩',
    '玢岩', '角斑岩', '细碧岩', '蛇绿岩', '科马提岩', '辉绿岩', '煌斑岩',
    # 火山碎屑岩
    '火山角砾岩', '集块岩', '熔结凝灰岩', '沉凝灰岩',
    # 其他
    '石英', '长石', '云母', '角闪石', '辉石', '方解石', '白云石',
    '燧石', '碧玉', '玉髓', '赤铁矿', '磁铁矿',
}

# 地质年代标准化
TIME_MAP = {
    '太古代': '太古代', '晚太古代': '太古代', '早太古代': '太古代',
    '元古代': '元古代', '古元古代': '古元古代', '中元古代': '中元古代',
    '新元古代': '新元古代',
    '前震旦纪': '前震旦纪', '震旦纪': '震旦纪',
    '早震旦世': '早震旦世', '晚震旦世': '晚震旦世',
    '南华纪': '南华纪', '早南华世': '南华纪',
    '寒武纪': '寒武纪', '奥陶纪': '奥陶纪', '志留纪': '志留纪',
    '泥盆纪': '泥盆纪', '石炭纪': '石炭纪', '二叠纪': '二叠纪',
    '三叠纪': '三叠纪', '侏罗纪': '侏罗纪', '白垩纪': '白垩纪',
    '古近纪': '古近纪', '新近纪': '新近纪', '第四纪': '第四纪',
    '早寒武世': '寒武纪', '中寒武世': '寒武纪', '晚寒武世': '寒武纪',
    '早奥陶世': '奥陶纪', '中奥陶世': '奥陶纪', '晚奥陶世': '奥陶纪',
    '早志留世': '志留纪', '中志留世': '志留纪', '晚志留世': '志留纪',
    '早泥盆世': '泥盆纪', '中泥盆世': '泥盆纪', '晚泥盆世': '泥盆纪',
    '早石炭世': '石炭纪', '中石炭世': '石炭纪', '晚石炭世': '石炭纪',
    '早二叠世': '二叠纪', '中二叠世': '二叠纪', '晚二叠世': '二叠纪',
    '早三叠世': '三叠纪', '中三叠世': '三叠纪', '晚三叠世': '三叠纪',
    '早侏罗世': '侏罗纪', '中侏罗世': '侏罗纪', '晚侏罗世': '侏罗纪',
    '早白垩世': '白垩纪', '中白垩世': '白垩纪', '晚白垩世': '白垩纪',
    '中元古代一纪': '中元古代', '中生代': '中生代',
    '印支期': '三叠纪', '燕山期': '侏罗纪-白垩纪', '喜马拉雅期': '新生代',
}

# ═══════════════════════════════════════════════════════════
# 提取函数
# ═══════════════════════════════════════════════════════════

def classify_entity(name):
    """根据地层名称判断实体类型"""
    name = name.strip()
    if name.endswith('群'):
        return 'GROUP'
    elif name.endswith('(岩)组') or name.endswith('岩组') or name.endswith('组'):
        return 'FORMATION'
    elif name.endswith('段'):
        return 'MEMBER'
    elif name.endswith('杂岩') or name.endswith('混杂岩'):
        return 'FORMATION'  # 按组处理
    return 'FORMATION'  # 默认按组处理

def extract_rocks(text):
    """从岩石组合描述中提取岩石类型名称"""
    if not text:
        return set()
    rocks = set()
    # 分词法：找所有2-4字组合，匹配已知岩石名
    text = text.replace('、', '，').replace('；', '，').replace(';', '，')
    # 直接用已知岩石名匹配
    for r in ROCK_NAMES:
        if r in text:
            rocks.add(r)
    # 补充：前缀+岩石名的组合 (如"黑云斜长变粒岩" -> "变粒岩"和"黑云斜长变粒岩")
    for r in list(rocks):
        idx = text.find(r)
        if idx >= 2:
            # 往前找修饰词
            prefix = text[max(0, idx-4):idx]
            prefix = prefix.strip('，。、 ')
            if len(prefix) >= 1 and not any(c in prefix for c in '，。、；;'):
                full_name = prefix + r
                if 2 <= len(full_name) <= 8:
                    rocks.add(full_name)
    return rocks

def extract_locations(text):
    """从分布描述中提取地理位置"""
    if not text:
        return set()
    locations = set()
    # 匹配地名+后缀模式
    patterns = [
        r'([一-鿿]{2,3}(?:市|县|区|镇|乡|村|地区|一带|等地|东部|西部|南部|北部|中部|东南|西南|东北|西北|沿海|内陆))',
    ]
    for pat in patterns:
        matches = re.findall(pat, text)
        for m in matches:
            if len(m) >= 2:
                locations.add(m)
    return locations

def parse_time(time_str):
    """标准化地质年代"""
    if not time_str:
        return None
    time_str = time_str.strip()
    if time_str in TIME_MAP:
        return TIME_MAP[time_str]
    # 模糊匹配
    for key, val in TIME_MAP.items():
        if key in time_str:
            return val
    return time_str  # 保留原始值

def extract_fossils(text):
    """从古生物组合提取化石"""
    if not text:
        return set()
    fossils = set()
    fossil_pats = [
        r'([一-鿿]{1,4}(?:类|虫|贝|螺|藻|珊瑚|蜓|菊石|腕足|三叶虫|笔石|牙形石|孢粉|介形虫|双壳类|腹足类|放射虫|蓝绿藻|迭层石))',
    ]
    for pat in fossil_pats:
        matches = re.findall(pat, text)
        fossils.update(matches)
    return fossils

# ═══════════════════════════════════════════════════════════
# 主转换流程
# ═══════════════════════════════════════════════════════════

def main():
    excel_path = r'E:\地质文本空间化平台\岩石地层-0312地大武汉.xlsx'
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    ws = wb['Sheet1']

    headers = [str(c.value) if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]

    # 提取所有行
    formations = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = {}
        for i, (h, v) in enumerate(zip(headers, row)):
            d[h] = str(v).strip() if v else ''
        formations.append(d)

    print(f'Loaded {len(formations)} stratigraphic units from Excel')

    G = nx.DiGraph()
    all_triples = []
    stats = Counter()

    for fm in formations:
        name = fm.get('地层名称', '')
        if not name:
            continue

        etype = classify_entity(name)
        clean_name = name.replace('(岩)', '').strip()
        G.add_node(clean_name, entity_type=etype)

        # 1. COMPOSED_OF triples (岩性组成)
        rock_text = fm.get('岩石组合特征', '')
        rocks = extract_rocks(rock_text)
        for rock in rocks:
            if rock != clean_name and len(rock) >= 2:
                G.add_node(rock, entity_type='LITHOLOGY')
                G.add_edge(clean_name, rock, relation='COMPOSED_OF')
                all_triples.append((clean_name, 'COMPOSED_OF', rock))
                stats['COMPOSED_OF'] += 1

        # 2. DATED_AS triples (地质年代)
        time_str = fm.get('地质年代', '')
        time_norm = parse_time(time_str)
        if time_norm:
            G.add_node(time_norm, entity_type='GEOLOGICAL_TIME')
            G.add_edge(clean_name, time_norm, relation='DATED_AS')
            all_triples.append((clean_name, 'DATED_AS', time_norm))
            stats['DATED_AS'] += 1

        # 3. DISTRIBUTED_IN triples (分布)
        dist_text = fm.get('分布', '')
        locations = extract_locations(dist_text)
        # 也加入省份
        province = fm.get('省份', '')
        if province and province != 'None':
            locations.add(province)
        for loc in locations:
            if len(loc) >= 2 and loc != clean_name:
                G.add_node(loc, entity_type='LOCATION')
                G.add_edge(clean_name, loc, relation='DISTRIBUTED_IN')
                all_triples.append((clean_name, 'DISTRIBUTED_IN', loc))
                stats['DISTRIBUTED_IN'] += 1

        # 4. SAME_AS triples (同物异名)
        synonym = fm.get('同物异名', '')
        if synonym and synonym != 'None':
            for syn in re.split(r'[、，,等]', synonym):
                syn = syn.strip()
                if syn and len(syn) >= 2 and syn != clean_name:
                    G.add_node(syn, entity_type=etype)
                    G.add_edge(clean_name, syn, relation='SAME_AS')
                    all_triples.append((clean_name, 'SAME_AS', syn))
                    stats['SAME_AS'] += 1

        # 5. CONTAINS_FOSSIL triples (化石)
        fossil_text = fm.get('古生物组合', '')
        fossils = extract_fossils(fossil_text)
        for fossil in fossils:
            if len(fossil) >= 2:
                G.add_node(fossil, entity_type='FOSSIL')
                G.add_edge(clean_name, fossil, relation='CONTAINS_FOSSIL')
                all_triples.append((clean_name, 'CONTAINS_FOSSIL', fossil))
                stats['CONTAINS_FOSSIL'] += 1

        # 6. 从创建地点提取SECTION
        loc_name = fm.get('创建地点', '')
        if loc_name and loc_name != 'None':
            section_name = f"{clean_name}层型剖面"
            G.add_node(section_name, entity_type='SECTION')
            G.add_node(loc_name, entity_type='LOCATION')
            G.add_edge(clean_name, section_name, relation='MEASURED_AT')
            G.add_edge(section_name, loc_name, relation='DISTRIBUTED_IN')
            stats['MEASURED_AT'] += 1

    # ═══════════════════════════════════════════════════════
    # 层级关系推断
    # ═══════════════════════════════════════════════════════
    # 从分段信息推断 BELONGS_TO (Member -> Formation)
    for fm in formations:
        name = fm.get('地层名称', '').replace('(岩)', '').strip()
        segment_info = fm.get('分段信息', '')
        if segment_info and segment_info != 'None':
            # 匹配"分为X段"或"A段、B段"等模式
            segs = re.findall(r'([一-鿿]{2,6}段)', segment_info)
            for seg in segs:
                G.add_node(seg, entity_type='MEMBER')
                G.add_edge(seg, name, relation='BELONGS_TO')
                G.add_edge(name, seg, relation='HAS_SUBUNIT')

    # ═══════════════════════════════════════════════════════
    # 保存
    # ═══════════════════════════════════════════════════════
    out_dir = os.path.join(os.path.dirname(__file__), '..', 'kb_storage', 'graphs')
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, 'experiment_kb.graphml')
    nx.write_graphml(G, out_path)

    print(f'\n=== KG Construction Complete ===')
    print(f'Nodes: {G.number_of_nodes()}')
    print(f'Edges: {G.number_of_edges()}')
    print(f'Triples: {len(all_triples)}')
    print(f'\nTriple type distribution:')
    for rel, count in stats.most_common():
        print(f'  {rel}: {count}')
    print(f'\nSaved to: {out_path}')

    # 保存三元组明细
    triple_path = os.path.join(out_dir, '..', 'experiment_triples.json')
    with open(triple_path, 'w', encoding='utf-8') as f:
        triple_data = [{'head': h, 'relation': r, 'tail': t} for h, r, t in all_triples]
        json.dump(triple_data, f, ensure_ascii=False, indent=2)
    print(f'Triples JSON: {triple_path}')

if __name__ == '__main__':
    main()
